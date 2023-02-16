import openpyxl
wb = openpyxl.load_workbook('store.xlsx')
print(wb.sheetnames)

for s in wb:
    print(s.title)

sheet = wb['Products']
sheet = wb.active

b2_cell = sheet['b2']
print(b2_cell.value)
print(sheet.cell(row=4, column=2).value)
print(sheet['a5'].data_type)

cell_range = sheet['b2:c11']
for p, pr in cell_range:
    print(f'prod: {p} \n price: {pr}')

print(f'Dimension: {sheet.dimensions}')
print(sheet.max_row, sheet.max_column)

# writing data
sheet['d2'] = 400121

## Adding a new row
new_product = (11, 'Tablet', 12, 600, 12 * 600)  # this is a tuple
sheet.append(new_product)

## Iterating and updating an entire column
for c, d, e in sheet['c2:e12']:
    e.value = c.value * d.value

## Saving the file
wb.save('store.xlsx')

# Creating new files
wb = openpyxl.Workbook()

## Getting the active sheet. This will be the only sheet in the workbook
sheet = wb.active

## Updating the new sheet using cells' addresses
sheet['A1'] = 'Year'
sheet['B1'] = 'Sales'

sales = {2017: 700000, 2018: 800000, 2019: 900000}  # dictionary

## Iterating over the dictionary and appending to the sheet
for k, v in sales.items():
    sheet.append((k, v))

## Saving the Workbook (in the same directory with the python script)
wb.save('sales.xlsx')


# formulas
wb = openpyxl.load_workbook('store.xlsx')

## Getting a sheet by name
sheet = wb['Products']

## Iterating over a cell range and writing an Excel formula in each cell of a column
for c, d, e in sheet['c2:e12']:
    e.value = f'={c.coordinate}*{d.coordinate}'
    # writing an Excel formula to each cell. The formula is written as a string

## Getting another sheet of the Workbook
sheet = wb['Sales 2018']

## Writing an Excel formula (=sum(B2:B13)
sheet['B14'] = '=sum(B2:B13)'

## Saving the file
wb.save('store.xlsx')

# sheet operations
sheet = wb['Products']

## Printing all sheet's methods
print(dir(sheet))

## Changing the sheet title
# sheet.title = 'Products for sale'


## Printing sheet properties
print(sheet.sheet_format)
print(sheet.sheet_properties)

## Create a new sheet on first position in the Workbook (by default will be the last)
wb.create_sheet('Turnover1', 0)

## Getting a sheet by name
sheet1 = wb['Turnover1']

## Removing a sheet by name
wb.remove(sheet1)

## Copying sheets in the same Workbook
source = wb['Turnover']
destination = wb.copy_worksheet(source)
# print(destination.title)

## Saving the changes
wb.save('store.xlsx')

# styles
my_cell = sheet['B4']

## Printing all available styles
# print(dir(openpyxl.styles))

## Changing and setting a new font
font = Font(name='Tahoma', size=16, color=colors.RED, bold=True, italic=True, strike=False)
my_cell.font = font

## Setting a Pattern Fill
fill = PatternFill(fill_type='solid', fgColor=colors.YELLOW)
my_cell.fill = fill

## Setting cell border
double_border_green = Side(border_style='double', color=colors.GREEN)
thin_border_red = Side(border_style='thin', color='FF0000')
my_cell.border = Border(left=double_border_green, right=thin_border_red, top=double_border_green,
                        bottom=thin_border_red)

## Setting cell's content alignment
alignment = Alignment(horizontal='right', vertical='center')
my_cell.alignment = alignment

## Copy styles from a cell to another
new_cell = sheet['B10']
new_font = copy(my_cell.font)
new_font.color = colors.GREEN
new_cell.font = new_font

## Saving the changes
wb.save('store.xlsx')