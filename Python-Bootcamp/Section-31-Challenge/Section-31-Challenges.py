# Challenge 1
import openpyxl, csv
from openpyxl.styles import *

wb = openpyxl.Workbook()
sheet = wb.active
rows = (
    ('Year', 'Sales'),
    (2017, 150000),
    (2018, 180000),
    (2019, 210000),
    (2020, 125000)
)
for r in rows:
    sheet.append(r)
# =====================================================================================================================
# Challenge 2
sheet = wb.active
sheet.title = 'COMPANY SALES'
wb.save('sales.xlsx')
# =====================================================================================================================
# Challenge 3
wb = openpyxl.load_workbook('sales_c3.xlsx')
sheet = wb.active
items = list()
for r in sheet.values:
    items.append(r)

vat = list()
for v in items[1:]:
    element = (v[0], v[1] * 0.15)
    vat.append(element)

wb.create_sheet('VAT')
sheet = wb['VAT']
sheet['A1'] = 'Year'
sheet['B1'] = 'VAT'

for r in vat:
    sheet.append(r)

wb.save('sales_and_vat.xlsx')
# =====================================================================================================================
# Challenge 4
wb = openpyxl.load_workbook('sales2_c4.xlsx')
sheet = wb.active
cell = sheet['B6']
cell.value = '=sum(B2:B5)'
cell = sheet['C6']
cell.value = 'Total Sales'
wb.save('sales2_c4_modified.xlsx')
# =====================================================================================================================
# Challenge 5
wb = openpyxl.load_workbook('sales2_c4_modified.xlsx')
sheet = wb.active
font = Font(name='Tahoma', size=16, color=colors.RED, bold=True, italic=False, strike=False)
cell_b6 = sheet['B6']
cell_b6.font = font
cell_c6 = sheet['c6']
cell_c6.font = font
wb.save('sales2_c4_modified_redText.xlsx')
# =====================================================================================================================
# Challenge 6
def csv2excel(csv_file, excel_file):
    # Load CSV file data into a list of lists
    with open(csv_file, 'r') as f:
        reader = csv.reader(f)
        data = list(reader)
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    # Write the data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook to the specified file
    wb.save(excel_file)
    print(f"{csv_file} exported to {excel_file} successfully!")
# =====================================================================================================================
# Challenge 7
def csv2excel(csv_file, excel_file, delim=','):
    # Load CSV file data into a list of lists
    with open(csv_file, 'r') as f:
        reader = csv.reader(f, delimiter=delim)
        data = list(reader)
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    # Write the data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook to the specified file
    wb.save(excel_file)
    print(f"{csv_file} exported to {excel_file} successfully!")
# =====================================================================================================================
# Challenge 8
def excel2csv(excel_file, csv_file):
    # Load the workbook and select the active worksheet
    wb = openpyxl.Workbook(filename=excel_file)
    ws = wb.active

    # Load the worksheet data into a list of lists
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    # Write the data to the CSV file
    with open(csv_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)

    print(f"{excel_file} exported to {csv_file} successfully!")